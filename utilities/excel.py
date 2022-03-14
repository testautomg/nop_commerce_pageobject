import openpyxl


def get_data(file_name, sheet_name):
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook[sheet_name]

    total_rows = sheet.max_row
    total_cols = sheet.max_column
    mainList = []

    for i in range(2, total_rows + 1):
        rowData = []
        for j in range(1, total_cols + 1):
            data = sheet.cell(row=i, column=j).value
            rowData.append(data)
        mainList.append(rowData)
    return mainList
